#!/usr/bin/env python3
"""
Process project images from media/projects into optimized WebP files
for the Ganson website. Reads project data from Excel, maps to image
folders, selects up to 8 images per project, resizes and compresses.
"""

import os
import re
import sys
import json

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

from PIL import Image, ImageOps, ImageDraw, ImageFont

# === Config ===
MEDIA_ROOT = "media/projects"
OUTPUT_ROOT = "public/assets/projects"
MAX_IMAGES = 8
MAX_WIDTH = 1920
WEBP_QUALITY = 80
MIN_FILE_SIZE = 10_000  # Skip files smaller than 10KB (likely thumbnails)

# === Slug generation ===
def slugify(name):
    s = name.lower().strip()
    s = re.sub(r"['''\u2019]", "", s)
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = s.strip("-")
    # collapse multiple hyphens
    s = re.sub(r"-+", "-", s)
    return s

# === Map Excel project names to media folder paths ===
# Each key is the Excel project name, value is list of folder paths to search
PROJECT_FOLDER_MAP = {
    "Poolbeg": ["01. Civils & Public Infrastucture/01. Poolbeg"],
    "Ardee Ambulance Centre": ["01. Civils & Public Infrastucture/03. Ardee Ambulance Centre"],
    "Edenderry Ambulance Station": ["01. Civils & Public Infrastucture/04. Edenderry Ambulance Station"],
    "HSE NDC Tullamore": ["01. Civils & Public Infrastucture/05. HSE NDC Tullamore"],
    "Creamatorium": ["01. Civils & Public Infrastucture/06. Creamatorium"],
    "Dundalk Driver Test Centre": ["01. Civils & Public Infrastucture/02. Dundalk Driver Test Centre"],
    "Ryanair Training Centre, \nCafe & Simulation Building": [
        "02. Commercial & Retail/01. Ryanair Training Centre, Cafe & Simulation Building"
    ],
    "Lidl Castlereagh": ["02. Commercial & Retail/02. Lidl Castlereagh"],
    "Home Bargins Lisburn": ["02. Commercial & Retail/04. Home Bargins Lisburn"],
    "Arden Road, Mixed Use Development, Tullamore": [
        "02. Commercial & Retail/07. Arden Road, Mixed Use Development, Tullamore",
        "05. Residential/05. Arden Road, Mixed Use Development, Tullamore",
    ],
    "The Keep": ["02. Commercial & Retail/09. The Keep, Belfast"],
    "St Lawrence O Toole's, Dublin": ["03. Education/01. St Lawrence O Toole_s, Dublin"],
    " St Lawrence O Toole's, Dublin": ["03. Education/01. St Lawrence O Toole_s, Dublin"],
    "Colaiste Chuchulainn Dundalk": ["03. Education/02. Colaiste Chuchulainn Dundalk"],
    "Grangegorman NS": [
        "03. Education/03. Grangegorman NS",
        "03. Grangegorman NS",
    ],
    "Lusk NS": ["03. Education/04. Lusk NS"],
    "Dun na Ri Kingscourt Cavan": ["03. Education/05. Dun na Ri Kingscourt Cavan"],
    "Naas Community College": ["03. Education/06. Naas Community College"],
    "O'Carolan College Nobber": ["03. Education/07. O_Carolan College Nobber"],
    "O\u2019Carolan College Nobber": ["03. Education/07. O_Carolan College Nobber"],
    " St Oliver Plunkett School Malahide": ["03. Education/08. St Oliver Plunkett School Malahide"],
    "St Oliver Plunkett School Malahide": ["03. Education/08. St Oliver Plunkett School Malahide"],
    "St. Michael's College, Dublin 4": ["03. Education/09. St. Michael_s College, Dublin 4"],
    "Stewarts School Lucan": ["03. Education/10. Stewarts School Lucan"],
    "Stephenstown Schools, Balbriggan": ["03. Education/11. Stephenstown Schools, Balbriggan"],
    "Scoil Phadraig Clane": ["03. Education/12. Scoil Phadraig Clane"],
    "Kolbe Special School": ["03. Education/13. Kolbe Special School"],
    "Kilbride Eductation Campus": ["03. Education/14. Kilbride Education Campus"],
    "Kilbride Education Campus": ["03. Education/14. Kilbride Education Campus"],
    "Ryevale Nursing Home": ["04. Healthcare/01. Ryevale Nursing Home"],
    "Carechoice Nursing Home, Trim": ["04. Healthcare/02. Carechoice Nursing Home, Trim"],
    "Carechoice Nursing Home, Marley": ["04. Healthcare/03. Carechoice Nursing Home, Marley"],
    "Carechoice Nursing Home, Parnell Road": ["04. Healthcare/04. Carechoice Nursing Home, Parnell Road"],
    "Carechoice Nursing Home, Swords": ["04. Healthcare/05. Carechoice Nursing Home, Swords"],
    "Dunshaughlin Nursing Home": [],
    "Moorehall Nursing Home": ["04. Healthcare/07. Moorehall Nursing Home"],
    "Kilbarrack Health Centre": ["04. Healthcare/08. Kilbarrack Health Centre"],
    "Lucan Health Centre": ["04. Healthcare/09. Lucan Health Centre"],
    "Irish Nursing Centre, Ballivor": ["04. Healthcare/10. Irish Nursing Centre, Ballivor"],
    "Donabate": ["05. Residential/01. Donabate Apartments"],
    "Circle Housing, Inchicore": ["05. Residential/08. Circle Housing, Inchicore"],
    "St. Mary's Mansions, Dublin 1": ["05. Residential/03. St. Mary_s Mansions, Dublin 1"],
    "Tooting Meadow, Scarlet Street, Drogheda": ["05. Residential/04. Scarlet Street, Drogheda"],
    "The Galleries, Donabate": ["05. Residential/06. The Galleries, Donabate"],
    "Arbour Hill Apartments": ["05. Residential/07. Arbour Hill Apartments"],
    "Room2 Hotel, Belfast": ["06. Hotel, Leisure & Student Residential/01. Room2 Hotel, Belfast"],
    "Hamilton Dock Hotel, Belfast": [],
    "Earls Hotel, Pembroke St, Dublin": ["06. Hotel, Leisure & Student Residential/03. Earls Hotel, Pembroke St, Dublin"],
    "Crumlin Hursing Home": [],
    "Crumlin Nursing Home": [],
    "Louth GAA Stadium": [],
    "Bakers Corner, Student Residential": [],
    "James McSweeney House, Apartments": [],
    " HSE NDC Tullamore": ["01. Civils & Public Infrastucture/05. HSE NDC Tullamore"],
}


def find_images(folder_paths, media_root=MEDIA_ROOT, min_size=MIN_FILE_SIZE):
    """Find image files, prioritising '02. Project Handover Photos' subfolders.
    Only falls back to other subfolders if no handover photos are found."""
    handover_images = []
    fallback_images = []

    for rel_path in folder_paths:
        full_path = os.path.join(media_root, rel_path)
        if not os.path.isdir(full_path):
            print(f"  WARNING: folder not found: {full_path}")
            continue

        # Look for handover photos subfolder (handles both correct and misspelled)
        handover_dir = None
        for entry in os.listdir(full_path):
            entry_path = os.path.join(full_path, entry)
            if os.path.isdir(entry_path) and entry.startswith("02.") and "roject" in entry:
                handover_dir = entry_path
                break

        if handover_dir:
            # Recursively find images in the handover folder (may have subfolders)
            for root, dirs, files in os.walk(handover_dir):
                for f in sorted(files):
                    ext = f.lower().rsplit('.', 1)[-1] if '.' in f else ''
                    if ext in ('jpg', 'jpeg', 'png', 'webp'):
                        file_path = os.path.join(root, f)
                        size = os.path.getsize(file_path)
                        if size >= min_size:
                            handover_images.append(file_path)

        # Also collect fallback images from all subfolders (in case handover is empty)
        for root, dirs, files in os.walk(full_path):
            for f in sorted(files):
                ext = f.lower().rsplit('.', 1)[-1] if '.' in f else ''
                if ext in ('jpg', 'jpeg', 'png', 'webp'):
                    file_path = os.path.join(root, f)
                    size = os.path.getsize(file_path)
                    if size >= min_size:
                        fallback_images.append(file_path)

    # Use handover images if available, otherwise fall back
    if handover_images:
        print(f"  Using {len(handover_images)} handover photos")
        return sorted(handover_images)
    elif fallback_images:
        print(f"  No handover photos — falling back to {len(fallback_images)} other images")
        return sorted(fallback_images)
    return []


def select_images(images, n=MAX_IMAGES):
    """Select n images spread evenly across the list."""
    if not images:
        return []
    if len(images) <= n:
        return images
    step = len(images) / n
    return [images[int(i * step)] for i in range(n)]


def optimize_image(src, dst, max_width=MAX_WIDTH, quality=WEBP_QUALITY):
    """Resize and convert image to WebP."""
    try:
        img = Image.open(src)
        # Handle EXIF rotation
        img = ImageOps.exif_transpose(img)
        # Convert to RGB
        if img.mode in ('RGBA', 'P', 'LA'):
            bg = Image.new('RGB', img.size, (255, 255, 255))
            if img.mode == 'P':
                img = img.convert('RGBA')
            if 'A' in img.mode:
                bg.paste(img, mask=img.split()[-1])
                img = bg
            else:
                img = img.convert('RGB')
        elif img.mode != 'RGB':
            img = img.convert('RGB')

        # Resize if wider than max
        w, h = img.size
        if w > max_width:
            ratio = max_width / w
            new_w = max_width
            new_h = int(h * ratio)
            img = img.resize((new_w, new_h), Image.LANCZOS)

        # Save as WebP
        os.makedirs(os.path.dirname(dst), exist_ok=True)
        img.save(dst, 'WEBP', quality=quality, method=4)
        return os.path.getsize(dst)
    except Exception as e:
        print(f"  ERROR processing {src}: {e}")
        return 0


def create_placeholder(dst, width=1920, height=1080):
    """Create a placeholder WebP image for projects without photos."""
    img = Image.new('RGB', (width, height), (18, 32, 71))  # Navy color
    draw = ImageDraw.Draw(img)
    # Draw subtle grid lines
    for x in range(0, width, 60):
        draw.line([(x, 0), (x, height)], fill=(26, 45, 94), width=1)
    for y in range(0, height, 60):
        draw.line([(0, y), (width, y)], fill=(26, 45, 94), width=1)
    # Draw "COMING SOON" text
    try:
        font = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", 48)
    except:
        font = ImageFont.load_default()
    text = "COMING SOON"
    bbox = draw.textbbox((0, 0), text, font=font)
    tw = bbox[2] - bbox[0]
    th = bbox[3] - bbox[1]
    draw.text(((width - tw) / 2, (height - th) / 2), text, fill=(208, 9, 24), font=font)
    os.makedirs(os.path.dirname(dst), exist_ok=True)
    img.save(dst, 'WEBP', quality=80)


def read_excel(path="media/Project Information.xlsx"):
    """Read project data from Excel."""
    wb = openpyxl.load_workbook(path)
    projects = []
    seen_names = set()

    for sheet_name in ["For Website", "To be reviewed"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or not row[1]:  # Skip empty rows
                continue
            photos = str(row[0]).strip() if row[0] else "No"
            name = str(row[1]).strip() if row[1] else ""
            if not name:
                continue

            # Skip duplicates (Arden Road appears in both Residential and Commercial)
            # Keep the first occurrence
            clean_name = re.sub(r'\s+', ' ', name).strip()
            if clean_name in seen_names:
                continue
            seen_names.add(clean_name)

            status = str(row[2]).strip() if row[2] else "Completed"
            sector = str(row[3]).strip() if row[3] else ""
            client = str(row[4]).strip() if row[4] else ""
            location = str(row[5]).strip() if row[5] else ""
            year = str(row[6]).strip() if row[6] else ""
            overview = str(row[7]).strip() if row[7] else ""

            # Clean year (remove .0)
            if year.endswith('.0'):
                year = year[:-2]

            # Map sector names
            sector_map = {
                "Civils & Public Infastructure": "civil-engineering",
                "Civils & Public Infrastructure": "civil-engineering",
                "Commercial": "commercial",
                "Education": "education",
                "Healthcare": "healthcare",
                "Residential": "residential",
                "Hotel, Leisure & Student Res": "hotel-leisure",
                "Hotel, Leisure & Student Residential": "hotel-leisure",
            }
            sector_slug = sector_map.get(sector, "commercial")

            # Map status
            status_slug = "current" if status.lower() == "current" else "completed"

            # Only include projects from "For Website" sheet,
            # plus Lidl Castlereagh from "To be reviewed" (has photos)
            if sheet_name == "To be reviewed" and photos.lower() != "yes":
                continue

            projects.append({
                "name": name,
                "photos_uploaded": photos.lower() == "yes",
                "status": status_slug,
                "sector": sector_slug,
                "client": client,
                "location": location,
                "year": year,
                "overview": overview,
                "sheet": sheet_name,
            })

    return projects


def get_short_description(overview, max_len=250):
    """Extract first paragraph as a short description."""
    if not overview or overview == "None":
        return "Project details coming soon."
    paragraphs = overview.split('\n\n')
    first = paragraphs[0].strip()
    if len(first) > max_len:
        # Cut at last sentence boundary before max_len
        cut = first[:max_len]
        last_period = cut.rfind('.')
        if last_period > 100:
            return cut[:last_period + 1]
        return cut + '...'
    return first


def main():
    print("=== Ganson Project Image Processor ===\n")

    # Read Excel
    print("Reading Excel...")
    projects = read_excel()
    print(f"Found {len(projects)} projects\n")

    results = []
    total_original = 0
    total_optimized = 0

    for proj in projects:
        name = proj["name"]
        slug = slugify(name)

        # Handle special slug cases
        if "ryanair" in slug and "training" in slug:
            slug = "ryanair-training-centre"
        if slug == "tooting-meadow-scarlet-street-drogheda":
            slug = "scarlet-street-drogheda"
        if slug == "crumlin-hursing-home":
            slug = "crumlin-nursing-home"
        if slug == "kilbride-eductation-campus":
            slug = "kilbride-education-campus"

        print(f"Processing: {name} → {slug}")

        # Find matching folder
        folder_key = name
        # Try exact match first, then variations
        folders = PROJECT_FOLDER_MAP.get(folder_key, None)
        if folders is None:
            # Try with leading space stripped
            folders = PROJECT_FOLDER_MAP.get(folder_key.strip(), None)
        if folders is None:
            # Try without newlines
            clean_key = re.sub(r'\s+', ' ', folder_key).strip()
            for k, v in PROJECT_FOLDER_MAP.items():
                clean_k = re.sub(r'\s+', ' ', k).strip()
                if clean_k == clean_key:
                    folders = v
                    break
        if folders is None:
            folders = []
            print(f"  No folder mapping found")

        # Find source images
        source_images = find_images(folders) if folders else []
        print(f"  Found {len(source_images)} source images")

        # Select up to MAX_IMAGES
        selected = select_images(source_images, MAX_IMAGES)
        has_photos = len(selected) > 0

        output_dir = os.path.join(OUTPUT_ROOT, slug)
        os.makedirs(output_dir, exist_ok=True)

        gallery_paths = []

        if has_photos:
            for i, src in enumerate(selected):
                dst = os.path.join(output_dir, f"{i+1:02d}.webp")
                orig_size = os.path.getsize(src)
                new_size = optimize_image(src, dst)
                total_original += orig_size
                total_optimized += new_size
                web_path = f"/assets/projects/{slug}/{i+1:02d}.webp"
                gallery_paths.append(web_path)
                print(f"  {os.path.basename(src)}: {orig_size/1024:.0f}KB → {new_size/1024:.0f}KB")
        else:
            # Create placeholder
            placeholder_path = os.path.join(output_dir, "placeholder.webp")
            create_placeholder(placeholder_path)
            web_path = f"/assets/projects/{slug}/placeholder.webp"
            gallery_paths.append(web_path)
            print(f"  Created placeholder image")

        results.append({
            "name": name,
            "slug": slug,
            "has_photos": has_photos,
            "image": gallery_paths[0] if gallery_paths else "",
            "gallery": gallery_paths,
            **{k: proj[k] for k in ["status", "sector", "client", "location", "year", "overview"]},
        })
        print()

    # Summary
    print(f"\n=== Summary ===")
    print(f"Projects processed: {len(results)}")
    print(f"With photos: {sum(1 for r in results if r['has_photos'])}")
    print(f"Placeholders: {sum(1 for r in results if not r['has_photos'])}")
    print(f"Original size: {total_original/1024/1024:.0f}MB")
    print(f"Optimized size: {total_optimized/1024/1024:.0f}MB")
    if total_original > 0:
        print(f"Compression ratio: {(1 - total_optimized/total_original)*100:.0f}% reduction")

    # Write results JSON for reference
    with open("scripts/project-image-results.json", "w") as f:
        json.dump(results, f, indent=2)
    print(f"\nResults written to scripts/project-image-results.json")


if __name__ == "__main__":
    main()
